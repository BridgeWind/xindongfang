import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.reader.excel import load_workbook
import os
#读操作读取文件
file_home = 'D:/Desktop/xdf/day5/义乌1 12.24高阳.xlsx'
workbook = xlrd.open_workbook(file_home)
sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
#写操作读取文件
wb=load_workbook(filename=file_home)
sheet_ranges=wb['Sheet1']
#print(sheet_ranges['G8'].value)

#获取指定day下非空的人名文件夹
filepath = 'D:/Desktop/xdf/day5'
del_dir = os.listdir(filepath)
dir_notempty_set = []
for f in del_dir:
	file_path = os.path.join(filepath, f)
	if os.path.isdir(file_path):
		sub_dir_files = os.listdir(file_path)
		if(len(sub_dir_files)>0):
			#今日已打卡人名
			dir_notempty_set.append(f)
print(dir_notempty_set)

row_counter = 1
for row in worksheet.get_rows():
	product_column = row[1]  # 学员标题所在的列
	product_value = product_column.value  # 项目名
	if product_value != '学员姓名' and product_value != '日期' and product_value != '群聊名称' and product_value != '校区' and product_value != '批改助教' and product_value != '':  # 排除第一行
		for daka_name in dir_notempty_set:
			if(daka_name == product_value):
				#对应天数打卡值入1	
				print(row_counter)
				sheet_ranges['G'+str(row_counter)] = 1
	row_counter = row_counter + 1

wb.save(file_home)
print(worksheet.cell(61,6))
